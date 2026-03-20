"""
reporter.py — Excel Report Generator
Produces professional .xlsx vulnerability reports with all 13 required fields.
"""

import uuid
import logging
from datetime import datetime, date
from pathlib import Path

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

import config
from scanner import ScanResult
from cve_lookup import CVERecord
from ai_analyst import AIAnalysis

logger = logging.getLogger(__name__)

# ── Column definitions (order matters — matches Excel schema) ──────────────────
COLUMNS = [
    ("ID",                 18),
    ("IP_Address",         16),
    ("Operating_System",   28),
    ("DNS_Name",           28),
    ("Status",             14),
    ("CVE_Name",           18),
    ("CVSS_Score",         12),
    ("Remediation_Target", 20),
    ("Summary",            55),
    ("Solution",           55),
    ("Public_Exploit",     16),
    ("Vulnerability",      35),
    ("Reference",          55),
]


def _severity_color(score: float) -> str:
    """Return hex fill color based on CVSS score."""
    if score >= 9.0:
        return config.EXCEL_CRITICAL_COLOR
    elif score >= 7.0:
        return config.EXCEL_HIGH_COLOR
    elif score >= 4.0:
        return config.EXCEL_MEDIUM_COLOR
    return config.EXCEL_LOW_COLOR


def _thin_border() -> Border:
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _style_header_row(ws, num_cols: int) -> None:
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(
            name="Calibri", bold=True, color="FFFFFF", size=11
        )
        cell.fill = PatternFill(
            fill_type="solid", fgColor=config.EXCEL_HEADER_COLOR
        )
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = _thin_border()


def _style_data_row(ws, row: int, cvss_score: float, num_cols: int) -> None:
    bg_color = "FFFFFF" if row % 2 == 0 else "F5F5F5"
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.font = Font(name="Calibri", size=10)
        cell.fill = PatternFill(fill_type="solid", fgColor=bg_color)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = _thin_border()

    # Highlight CVSS cell with severity color
    cvss_cell = ws.cell(row=row, column=7)  # Column G = CVSS_Score
    cvss_cell.fill = PatternFill(
        fill_type="solid", fgColor=_severity_color(cvss_score)
    )
    cvss_cell.font = Font(
        name="Calibri", bold=True, size=10,
        color="FFFFFF" if cvss_score >= 7.0 else "000000"
    )
    cvss_cell.alignment = Alignment(horizontal="center", vertical="top")


def _build_row(
    scan: ScanResult,
    cve: CVERecord,
    analysis: AIAnalysis,
    status: str = "Open",
) -> dict:
    """Build a single data row as an ordered dict matching COLUMNS."""
    reference = cve.references[0] if cve.references else f"https://nvd.nist.gov/vuln/detail/{cve.cve_id}"
    return {
        "ID":                 str(uuid.uuid4())[:8].upper(),
        "IP_Address":         scan.ip,
        "Operating_System":   scan.os,
        "DNS_Name":           scan.dns_name or "(none)",
        "Status":             status,
        "CVE_Name":           cve.cve_id,
        "CVSS_Score":         cve.cvss_score,
        "Remediation_Target": analysis.remediation_target.strftime("%Y-%m-%d"),
        "Summary":            analysis.summary,
        "Solution":           analysis.solution,
        "Public_Exploit":     "Yes" if cve.public_exploit else "No",
        "Vulnerability":      f"{cve.related_service.upper()} (port {cve.related_port}) — {cve.cvss_severity.upper()}",
        "Reference":          reference,
    }


def generate_report(
    data: list[tuple[ScanResult, CVERecord, AIAnalysis]],
    output_path: Path | None = None,
) -> Path:
    """
    Generate an Excel vulnerability report.

    Args:
        data: list of (ScanResult, CVERecord, AIAnalysis) tuples
        output_path: override default report path

    Returns:
        Path to generated .xlsx file
    """
    if not output_path:
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
        output_path = config.REPORTS_DIR / f"vuln_report_{timestamp}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vulnerabilities"

    # ── Title row ─────────────────────────────────────────────────────────────
    title_cell = ws.cell(row=1, column=1, value="VulnAgent — Vulnerability Report")
    ws.merge_cells(f"A1:{get_column_letter(len(COLUMNS))}1")
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill(fill_type="solid", fgColor="1F3864")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Sub-header: scan metadata
    meta = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  Targets: {', '.join(config.TARGETS)}  |  Total Findings: {len(data)}"
    meta_cell = ws.cell(row=2, column=1, value=meta)
    ws.merge_cells(f"A2:{get_column_letter(len(COLUMNS))}2")
    meta_cell.font = Font(name="Calibri", italic=True, size=9, color="666666")
    meta_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # ── Header row ────────────────────────────────────────────────────────────
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name.replace("_", " "))
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    _style_header_row(ws, len(COLUMNS))
    ws.row_dimensions[3].height = 22

    # ── Data rows ─────────────────────────────────────────────────────────────
    if not data:
        empty_cell = ws.cell(row=4, column=1, value="No vulnerabilities found in this scan.")
        ws.merge_cells(f"A4:{get_column_letter(len(COLUMNS))}4")
        empty_cell.alignment = Alignment(horizontal="center", vertical="center")
    else:
        for row_idx, (scan, cve, analysis) in enumerate(data, start=4):
            row_data = _build_row(scan, cve, analysis)
            for col_idx, (col_name, _) in enumerate(COLUMNS, start=1):
                ws.cell(row=row_idx, column=col_idx, value=row_data[col_name])
            _style_data_row(ws, row_idx, cve.cvss_score, len(COLUMNS))
            ws.row_dimensions[row_idx].height = 80  # tall for wrapped text

    # ── Freeze panes & filter ─────────────────────────────────────────────────
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(COLUMNS))}3"

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    _write_summary_sheet(ws2, data)

    # ── Save ──────────────────────────────────────────────────────────────────
    # Write to temp file then rename to avoid Excel file-lock issues
    tmp_path = output_path.with_suffix(".tmp.xlsx")
    wb.save(tmp_path)
    tmp_path.rename(output_path)

    logger.info(f"Report saved: {output_path}")
    return output_path


def _write_summary_sheet(ws, data: list) -> None:
    """Write a summary statistics sheet."""
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15

    critical = sum(1 for _, c, _ in data if c.cvss_score >= 9.0)
    high     = sum(1 for _, c, _ in data if 7.0 <= c.cvss_score < 9.0)
    medium   = sum(1 for _, c, _ in data if 4.0 <= c.cvss_score < 7.0)
    low      = sum(1 for _, c, _ in data if 0 < c.cvss_score < 4.0)
    exploitable = sum(1 for _, c, _ in data if c.public_exploit)

    rows = [
        ("SCAN SUMMARY", ""),
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Targets Scanned", ", ".join(config.TARGETS)),
        ("Total Findings", len(data)),
        ("", ""),
        ("SEVERITY BREAKDOWN", ""),
        ("Critical (CVSS 9-10)", critical),
        ("High (CVSS 7-8.9)",    high),
        ("Medium (CVSS 4-6.9)",  medium),
        ("Low (CVSS 0.1-3.9)",   low),
        ("", ""),
        ("With Public Exploit",  exploitable),
    ]

    for row_idx, (label, value) in enumerate(rows, start=1):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=value)
        if not value:  # section headers
            ws.cell(row=row_idx, column=1).font = Font(bold=True)


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    print("Reporter module loaded. Run agent.py to generate a report.")
